"""
Audit tool: compare data/profit_pool_all_industries_2001-2025.csv against the
original Google Sheet it was compiled from (one tab per year, in Damodaran's
raw column layout) to check for data-entry errors before trusting the CSV.

Not part of the regular clean_data.py -> export_slim.py -> build_standalone.py
pipeline -- run this by hand whenever the source sheet is updated (e.g. a new
year is added) and you want to re-verify the CSV against it.

Usage:
    1. Export the Google Sheet as .xlsx (File > Download > Microsoft Excel,
       or via the Drive API with exportMimeType
       application/vnd.openxmlformats-officedocument.spreadsheetml.sheet --
       plain CSV export only returns one tab, so it has to be xlsx to get
       all 25 year-tabs at once) and save it as source_sheet.xlsx next to
       this script (i.e. src/source_sheet.xlsx -- gitignored, not committed).
    2. cd src && python3 compare_with_source_sheet.py
    3. Read the printed summary and src/diff_report.json (also gitignored)
       for the details.

What it checks: for every (year, industry) row in the CSV, is there a
matching row in the sheet with the same NumFirms/ROIC/WACC/spread/capital/EVA
(within float-rounding tolerance)? It also reports which (year, industry)
combinations exist in one source but not the other.

Last run (2026-09-05): 2,348 common rows, 0 numeric mismatches. 112 rows
existed only in the sheet, all of them industries whose ROIC was unusable
that year in the source itself (literal "NA", or a formula error like
#VALUE!/#DIV/0!/#####) -- consistent with the CSV's existing policy of
dropping rows it can't compute a spread for, not a copy error. A further 29
sheet-only rows were footer noise (aggregate "Grand Total" rows, and for
2019-2023 a block of column-definition rows) rather than industries, and
don't belong in either file.
"""
import csv
import json
import os
import sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "source_sheet.xlsx")
CSV = os.path.join(HERE, "..", "data", "profit_pool_all_industries_2001-2025.csv")
REPORT = os.path.join(HERE, "diff_report.json")


def parse_sheet():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    records = {}  # (year, industry) -> dict
    for year_str in wb.sheetnames:
        year = int(year_str)
        ws = wb[year_str]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            name = row[0]
            if name is None:
                continue
            name = str(name).strip()
            if name.startswith("Total Market") or name == "Grand Total":
                continue
            if not isinstance(row[1], (int, float)):
                # footer/definition rows (e.g. the 2019-2023 "ROC"/"EVA" blocks)
                continue

            def clean(v):
                if v is None or v in ("NA", "#VALUE!", "#DIV/0!", "#####"):
                    return None
                if isinstance(v, str):
                    v = v.strip().replace(",", "").replace("$", "")
                    is_pct = v.endswith("%")
                    v = v.rstrip("%")
                    try:
                        num = float(v)
                    except ValueError:
                        return None
                    return num / 100 if is_pct else num
                return v

            numfirms, roic, wacc, spread, capital, eva = (clean(x) for x in row[1:7])

            records[(year, name)] = dict(
                numfirms=numfirms,
                roic_pct=round(roic * 100, 2) if roic is not None else None,
                wacc_pct=round(wacc * 100, 2) if wacc is not None else None,
                spread_pct=round(spread * 100, 2) if spread is not None else None,
                capital_mn=round(capital, 2) if capital is not None else None,
                eva_mn=round(eva, 2) if eva is not None else None,
            )
    return records


def parse_csv():
    records = {}
    with open(CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            year = int(float(r["Year"]))
            name = r["Industry"].strip()

            def num(key):
                v = r.get(key, "")
                if v is None or v == "":
                    return None
                try:
                    return float(v)
                except ValueError:
                    return None

            records[(year, name)] = dict(
                numfirms=num("NumFirms"),
                roic_pct=num("ROIC_pct"),
                wacc_pct=num("WACC_pct"),
                spread_pct=num("ROIC_minus_WACC_pct"),
                capital_mn=num("Invested_Capital_USD_mn"),
                eva_mn=num("Economic_Profit_USD_mn"),
            )
    return records


def close(a, b, tol):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"Missing {XLSX} -- see this script's docstring for how to export it.")

    sheet = parse_sheet()
    csv_data = parse_csv()

    sheet_keys = set(sheet.keys())
    csv_keys = set(csv_data.keys())

    only_in_sheet = sorted(sheet_keys - csv_keys)
    only_in_csv = sorted(csv_keys - sheet_keys)
    common = sorted(sheet_keys & csv_keys)

    field_tol = dict(numfirms=0.5, roic_pct=0.06, wacc_pct=0.06, spread_pct=0.06)

    mismatches = []
    for key in common:
        s, c = sheet[key], csv_data[key]
        diffs = {}
        for field in ["numfirms", "roic_pct", "wacc_pct", "spread_pct", "capital_mn", "eva_mn"]:
            sv, cv = s[field], c[field]
            if field in ("capital_mn", "eva_mn"):
                if sv is None and cv is None:
                    ok = True
                elif sv is None or cv is None:
                    ok = False
                else:
                    ok = abs(sv - cv) <= max(0.5, abs(sv) * 0.005)
            else:
                ok = close(sv, cv, field_tol[field])
            if not ok:
                diffs[field] = (sv, cv)
        if diffs:
            mismatches.append((key, diffs))

    print(f"Common (year, industry) rows: {len(common)}")
    print(f"Numeric mismatches: {len(mismatches)}")
    print(f"Only in sheet (missing from CSV): {len(only_in_sheet)}")
    print(f"Only in CSV (not in sheet): {len(only_in_csv)}")

    with open(REPORT, "w") as f:
        json.dump(
            dict(
                only_in_sheet=[list(k) for k in only_in_sheet],
                only_in_csv=[list(k) for k in only_in_csv],
                mismatches=[[list(k), d] for k, d in mismatches],
            ),
            f, indent=2, default=str,
        )
    print(f"Wrote {REPORT}")


if __name__ == "__main__":
    main()
